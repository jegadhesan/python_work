import pandas as pd
import numpy as np
import os as os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# Load the workbooks Src_2023
workbook1 = load_workbook('File1.xlsx', data_only=True)
workbook2 = load_workbook('File2.xlsx', data_only=True)


# Function to check if a row or column is empty
def is_empty(cells):
    return all(cell is None or cell == '' for cell in cells)


# Function to extract data blocks from a sheet
def extract_data_blocks(sheet):
    data_blocks = []
    current_block = []

    # Extract data blocks separated by empty rows
    for row in sheet.iter_rows(values_only=True):
        if is_empty(row):  # Check for empty row
            if current_block:
                data_blocks.append(current_block)
                current_block = []
        else:
            current_block.append(row)

    if current_block:  # Add the last block if it exists
        data_blocks.append(current_block)

    # Handle empty columns within each block
    final_blocks = []

    for block in data_blocks:
        transposed_block = list(zip(*block))
        current_block = []
        for col_index, col in enumerate(transposed_block):
            if is_empty(col):  # Check for empty column
                if current_block:
                    final_blocks.append(list(zip(*current_block)))  # Transpose back
                    current_block = []
            else:
                # Check if the column name is "Percent"
                if block[0][col_index] in ("Percent", "Percent of Total Claims", "Percent of Total"):
                    # Round off and convert to percentage
                    rounded_col = [
                        (f"{round(cell * 100)}%" if isinstance(cell, (int, float)) else cell)
                        for cell in col
                    ]
                else:
                    rounded_col = col
                current_block.append(rounded_col)
        if current_block:
            final_blocks.append(list(zip(*current_block)))  # Transpose back

    return final_blocks

# Extract data blocks from both workbooks
dataframes_wb1 = {}
dataframes_wb2 = {}

#print(workbook1.sheetnames)
# if sheet_name.startswith("Cost"):
#   continue  # Skip processing for sheet names starting with "Cost"
blocks_to_skip = [0, 1, 3, 5]
# blocks_to_skip2 = [0, 1, 2, 3, 4, 5, 6, 7, 8]

for sheet_name in workbook1.sheetnames:
    if sheet_name in ["Member", "Member EffDates", "Member EndDates", "MemUserDef",\
                      "Subscriber","Provider", "Provider ProvSpec", "ServiceMed", \
                      "ProvSpec_Volume", "ServiceMed POS", "DiagProc Check", \
                      "ServiceRx", "Cost1Amt by Month", "ICD10 Diagnostic Codes", \
                     "ICD10 Procedure Codes"]:
        sheet = workbook1[sheet_name]
        # print(workbook1[sheet_name])
        data_blocks = extract_data_blocks(sheet)
        # print(f"{sheet_name}")
        for i, block in enumerate(data_blocks):
            if i in blocks_to_skip:  # Assuming you want to skip the first block
                continue
            headers = block[0]
            data = block[1:]  # Exclude headers from data
            df = pd.DataFrame(data, columns=headers)
            sheet_name = sheet_name.replace(' ', '')
            dataframes_wb1[f"{sheet_name}_block_{i}"] = df
            # print(f"{sheet_name}")
            print(f"{sheet_name}_wb1_block_{i}")

for sheet_name in workbook2.sheetnames:
    if sheet_name in ["Member", "Member EffDates", "Member EndDates", "MemUserDef",\
                      "Subscriber","Provider", "Provider ProvSpec", "ServiceMed", \
                      "ProvSpec_Volume", "ServiceMed POS", "DiagProc Check", \
                      "ServiceRx", "Cost1Amt by Month", "ICD10 Diagnostic Codes", \
                     "ICD10 Procedure Codes"]:
        sheet = workbook2[sheet_name]
        data_blocks = extract_data_blocks(sheet)
        # print(f"{sheet_name}")
        for i, block in enumerate(data_blocks):
            if i in blocks_to_skip:  # Assuming you want to skip the first block
                continue
            headers = block[0]
            data = block[1:]  # Exclude headers from data
            df = pd.DataFrame(data, columns=headers)
            sheet_name = sheet_name.replace(' ', '')
            dataframes_wb2[f"{sheet_name}_block_{i}"] = df
            # print(f"{sheet_name}")
            print(f"{sheet_name}_wb2_block_{i}")

print("Sheets Imported")


# Function to ensure headers are correctly set
def ensure_headers(df):
    if df.columns[0] is None or isinstance(df.columns[0], int):
        df.columns = [f"Column_{i}" for i in range(len(df.columns))]
    else:
        # Ensure all columns have valid names
        df.columns = [f"Column_{i}" if col is None else col for i, col in enumerate(df.columns)]
    return df


# Function to calculate percentage difference
def calculate_percentage_difference(row, col_x, col_y):
    try:
        val_x = row[col_x]
        val_y = row[col_y]
        # print(f"Processing row: {row.name}, {col_x}: {val_x}, {col_y}: {val_y}")  # Debugging output
        if pd.isna(val_x) or pd.isna(val_y):
            return None
        if not (isinstance(val_x, (int, float)) and isinstance(val_y, (int, float))):
            # print(f"Non-numeric data found: {val_x}, {val_y}")  # Debugging output
            return None  # Skip calculation if values are not numeric
        if val_x == 0:
            # print(f"Zero value found in denominator: {val_x}")  # Debugging output
            return None  # Avoid division by zero
        percentage_change = round(((val_y - val_x) / val_x) * 100)
        return f"({percentage_change}%)" if percentage_change < 0 else f"{percentage_change}%"
    except ZeroDivisionError:
        return None


def calculate_percentage_difference_pmpm(row, col_x, col_y):
    try:
        val_x = row[col_x]
        val_y = row[col_y]
        print(f"Processing row: {row.name}, {col_x}: {val_x}, {col_y}: {val_y}")  # Debugging output
        if pd.isna(val_x) or pd.isna(val_y):
            return None
        if not (isinstance(val_x, (int, float)) and isinstance(val_y, (int, float))):
            print(f"Non-numeric data found: {val_x}, {val_y}")  # Debugging output
            return None  # Skip calculation if values are not numeric
        if val_x == 0:
            print(f"Zero value found in denominator: {val_x}")  # Debugging output
            return None  # Avoid division by zero
        percentage_change = round(((val_y - val_x) / val_x) * 100)
        return f"({percentage_change}%)" if percentage_change < 0 else f"{percentage_change}%"
    except ZeroDivisionError:
        return None


# Function to calculate the difference between percent_x and percent_y
def calculate_difference(row):
    try:
        # val_x = row['Percent_PrevYear']
        # val_y = row['Percent_CurrYear']

        prev_col = [col for col in row.index if 'Percent' in col and 'Prev' in col][0]
        curr_col = [col for col in row.index if 'Percent' in col and 'Curr' in col][0]

        val_x = row[prev_col]
        val_y = row[curr_col]

        # print(f"val_x: {val_x}, val_y: {val_y}")  # Debugging output
        if pd.isna(val_x) or pd.isna(val_y):
            return None
        if isinstance(val_x, str):
            val_x = float(val_x.strip('%')) / 100
        if isinstance(val_y, str):
            val_y = float(val_y.strip('%')) / 100
        if not (isinstance(val_x, (int, float)) and isinstance(val_y, (int, float))):
            return None  # Skip calculation if values are not numeric
        if val_x == 0 and val_y == 0:
            return None
        diff = round((val_y - val_x) * 100, 2)  # Calculate the difference and round to 2 decimal places
        # print(f"val_x: {val_x}, val_y: {val_y}, Difference: {diff}")  # Debugging output
        # print(f"Difference: {diff}")  # Debugging output
        if diff >= 0:
            return f"{diff}%"
        elif diff < 0:
            return f"{abs(diff)}%"
        else:
            return None
    except Exception as e:
        print(f"Error in calculate_difference: {e}")
        return None


# Function to add block name prefix to the first column values
def add_block_name_prefix_to_first_column(df, block_name):
    # print(block_name)
    df.columns = [col + '_' + block_name if i == 0 else col for i, col in enumerate(df.columns)]
    # print(col + '_' + block_name)
    return df


# Function to extract the total value from Member_block_1
def get_total_value(dataframes, key, check_type, colname1, colname2):
    df = dataframes.get(key)
    if df is not None:
        # print(f"DataFrame for key '{key,dataframes, key, check_type, colname1, colname2}',  found.")
        # print(df)  # Print the DataFrame to check its contents
        total_value = df.loc[df[colname1] == check_type, colname2].sum()
        # print(f"Filtered DataFrame:\n{total_value}")  # Print the filtered DataFrame
        # print(total_value)  # This should be before the return statement if you want to see the output
        return total_value
    return None


# Member Sheet#
mem_cnt_x = get_total_value(dataframes_wb1, 'Member_block_1', 'Total Number of Members', 'Check Type', 'Value')
mem_cnt_y = get_total_value(dataframes_wb2, 'Member_block_1', 'Total Number of Members', 'Check Type', 'Value')

tmem_cnt_x2 = get_total_value(dataframes_wb1, 'Member_block_1', 'Total Number of Members', 'Check Type', 'Value')
tmem_cnt_x1 = get_total_value(dataframes_wb1, 'Member_block_1', 'Total Member Months*', 'Check Type', 'Value')

tmem_cnt_y2 = get_total_value(dataframes_wb2, 'Member_block_1', 'Total Number of Members', 'Check Type', 'Value')
tmem_cnt_y1 = get_total_value(dataframes_wb2, 'Member_block_1', 'Total Member Months*', 'Check Type', 'Value')

mem_cnt_val = get_total_value(dataframes_wb2, 'Member_block_1', 'Total Number of Members', 'Check Type', 'Value')
sub_cnt_val = get_total_value(dataframes_wb2, 'Subscriber_block_1', 'Total Number of Subscribers', 'Check Type',
                              'Value')
ser_cnt_val = get_total_value(dataframes_wb2, 'ServiceMed_block_1', 'Total Number of Members', 'Check Type', 'Count')
rx_cnt_val = get_total_value(dataframes_wb2, 'ServiceRx_block_1', 'Total Number of Members', 'Check Type', 'Count')

# print(tmem_cnt_x2)
# print(f"members_cnt_x: {tmem_cnt_x1,tmem_cnt_x2,tmem_cnt_y1,tmem_cnt_y2}, members_cnt_y: {members_cnt_y}")  # Debugging output

# print(f"Check : {tmem_cnt_x1, tmem_cnt_x2}, members_cnt_y: {tmem_cnt_y1, tmem_cnt_y2}")  # Debugging output

# Merge related DataFrames based on the first column
merged_dataframes = {}

# total_members = get_total_value(merged_dataframes, 'Member_block_1', 'Total Number of Members')
# print(total_members)

for key in dataframes_wb1:
    if key in dataframes_wb2:
        df1 = ensure_headers(dataframes_wb1[key])
        df2 = ensure_headers(dataframes_wb2[key])

        # Check if DataFrames have at least two columns
        if len(df1.columns) > 1 and len(df2.columns) > 1:
            merged_df = pd.merge(df1, df2, on=df1.columns[0], how='outer', suffixes=('_PrevYear', '_CurrYear'))

            # if key in ['DiagProcCheck_block_2')
            #   merged_df = pd.merge(df1, df2, on=df1.columns[0], how='outer', suffixes=('_PrevYear', '_CurrYear'))

            # Calculate percentage difference for the second column
            second_col_x = merged_df.columns[1]
            second_col_y = merged_df.columns[len(df1.columns)]

            # print(f"second_col_x: {second_col_x}, second_col_y: {second_col_y}")  # Debugging output
            merged_df['Val_diff'] = merged_df.apply(
                lambda row: calculate_percentage_difference(row, second_col_x, second_col_y), axis=1)

            # Add running total columns for Count_PrevYear
            if key in ['MemberEffDates_block_1', 'MemberEndDates_block_1', 'ProviderProvSpec_block_1' \
                    , 'Member_block_3', 'Member_block_4', 'Member_block_6', 'Member_block_9' \
                    , 'Provider_block_2', 'Provider_block_4', 'ProvSpec_Volume_block_2']:
                merged_df['Percentage_diff'] = merged_df.apply(calculate_difference, axis=1)

            if key in ['ServiceMedPOS_block_1', 'ServiceRx_block_1' \
                                                'ServiceRx_block_4', 'ServiceRx_block_5', 'ServiceRx_block_9',
                       'ServiceRx_block_11']:
                merged_df['Percentage_diff'] = merged_df.apply(calculate_difference, axis=1)

            if key in ['Cost1AmtbyMonth_block_1']:
                sixth_column_x = merged_df.columns[5]
                twelth_col_x = merged_df.columns[12]
                merged_df['Percentage_diff'] = merged_df.apply(
                    lambda row: calculate_percentage_difference(row, sixth_column_x, twelth_col_x), axis=1)
                # merged_df['Percentage_diff'] = merged_df.apply(calculate_percentage_difference_pmpm(row, sixth_column_x, twelth_col_x), axis=1)

            merged_df = add_block_name_prefix_to_first_column(merged_df, key)

            # Drop columns related to _PrevYear
            columns_to_drop = [col for col in merged_df.columns if '_PrevYear' in col]
            merged_df.drop(columns=columns_to_drop, inplace=True)

            # Drop Percent_CurrYear and Percent_PrevYear columns if they exist
            # merged_df = merged_df.drop(columns=['Percent_CurrYear', 'Percent_PrevYear'], errors='ignore')

            merged_dataframes[key] = merged_df

        else:
            None  # print(f"DataFrame for key {key} does not have enough columns")

# merged_dataframes
# merged_dataframes['ProvSpec_Volume_block_2']
print('Merged Dataframes Ready')

def compare_and_color_code(sheet, df, start_row, key):
    blue_fill = PatternFill(start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    purple_fill = PatternFill(start_color="b69cd9", end_color="b69cd9", fill_type="solid")  # Purple fill
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    mismatch_count = 0

    # Write the headers
    for col_idx, header in enumerate(df.columns, 1):
        cell = sheet.cell(row=start_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border

    # Write the data and apply color coding
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            col_name = df.columns[c_idx - 1]

            if key == 'Member_block_1':
                check_type = df.loc[r_idx - start_row - 1, 'Check Type_Member_block_1']
                if check_type in ['Total Member Months*', 'Total Number of Members']:
                    if col_name == 'Value_CurrYear':
                        value_total_member_months = df.loc[df['Check Type_Member_block_1'] \
                                                           == 'Total Member Months*', 'Value_PrevYear'].values
                        value_total_number_of_members = df.loc[df['Check Type_Member_block_1'] \
                                                               == 'Total Number of Members', 'Value_PrevYear'].values

                        if value_total_number_of_members > value_total_member_months:
                            cell.fill = blue_fill  # compare Total Member Months value with Total Number of Members

            if key == 'Member_block_1':
                if df.loc[r_idx - start_row - 1, 'Check Type_Member_block_1'] == '# Members with <1 Member Month*':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Value_PrevYear', 'Value_CurrYear']:
                        # Highlight non-zero values
                        if value != 0:
                            cell.fill = purple_fill

            if key == 'MemUserDef_block_1':
                if col_name in ['Mem_UserDef_1_PrevYear', 'Mem_UserDef_1_CurrYear']:
                    # Highlight non-zero values
                    if value != 'COM':
                        cell.fill = purple_fill

            if col_name == 'Percentage_diff':
                # print(col_name)
                if value is not None:
                    # Remove parentheses and strip the percentage sign
                    cleaned_value = value.replace('(', '').replace(')', '').strip('%')
                    per_diff_value = float(cleaned_value)
                    if per_diff_value > 10:
                        cell.fill = orange_fill

            # Apply color coding based on Val_diff and Per_diff
            if col_name == 'Val_diff' and isinstance(value, str):
                try:
                    percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                    if percentage_increase > 10:
                        cell.fill = blue_fill
                except ValueError as e:
                    print(f"Error converting Val_diff to float: {e}")

            # Apply color coding based on Val_diff and Per_diff
            if col_name == 'PMPM_diff' and isinstance(value, str):
                try:
                    percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                    if percentage_increase > 10:
                        cell.fill = blue_fill
                except ValueError as e:
                    print(f"Error converting Val_diff to float: {e}")

            if key in ('MemberEffDates_block_1', 'MemberEndDates_block_1'):
                if col_name == 'Percentage_diff':
                    if value is not None:
                        # Remove parentheses and strip the percentage sign
                        cleaned_value = value.replace('(', '').replace(')', '').strip('%')
                        try:
                            per_diff_value = float(cleaned_value)
                            # print(f"Processed value: {per_diff_value}")  # Debug print
                            if per_diff_value <= 0:
                                cell.fill = orange_fill
                                # print(f"Processed value: {per_diff_value}")  # Debug print
                        except ValueError:
                            print(f"Invalid value: {cleaned_value}")  # Debug print
                            pass
                            # Handle the case where cleaned_value is not a valid float

            if key == 'Member_block_12':
                if col_name in ['Record Count_PrevYear', 'Record Count_CurrYear']:
                    if isinstance(value, (int, float)) and value != 0:
                        cell.fill = purple_fill  # Highlight non-zero values

            if key == 'Member_block_18':
                if col_name in ['Record Count_PrevYear', 'Record Count_CurrYear']:
                    if isinstance(value, (int, float)) and value != 0:
                        cell.fill = purple_fill  # Highlight non-zero

            if key == 'ServiceRx_block_1':
                if df.loc[r_idx - start_row - 1, 'Check Type_ServiceRx_block_1'] == 'Total Number of Members':
                    if col_name in ['Count_CurrYear']:
                        # print(col_name)
                        # print(mem_cnt_val,sub_cnt_val)
                        # print('In')
                        mem_cnt_val != rx_cnt_val
                        cell.fill = purple_fill

            if key == 'ServiceRx_block_1':
                if df.loc[r_idx - start_row - 1, 'Check Type_ServiceRx_block_1'] == '# Pseudo Records':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Count_PrevYear', 'Count_CurrYear']:
                        # Highlight non-zero values
                        if value != 0:
                            cell.fill = purple_fill

            if key == 'ServiceRx_block_2':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceRx_block_2'] == '# Members without Eligibility Record':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceRx_block_3':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceRx_block_3'] == '# Records with DOS within Report Period':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 100:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceRx_block_3':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceRx_block_3'] == '# Records with DOS within Report Period':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 100:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceRx_block_10':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceRx_block_10'] == '# Records with Negative Allowed Amount':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceRx_block_13':
                if df.loc[
                    r_idx - start_row - 1, 'Check for Length Issues_ServiceRx_block_13'] == '# Records with Member ID > 32':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase is None or percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceRx_block_13':
                if df.loc[
                    r_idx - start_row - 1, 'Check for Length Issues_ServiceRx_block_13'] == '# Records with NDC < 11':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase is None or percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceRx_block_12':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceRx_block_12'] == '# Records with duplicate Unique Record ID ':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            print(percentage_increase)
                            if percentage_increase is None or percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_1':
                if df.loc[r_idx - start_row - 1, 'Check Type_ServiceMed_block_1'] == '# Pseudo Records':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Count_PrevYear', 'Count_CurrYear']:
                        # Highlight non-zero values
                        if value != 0:
                            cell.fill = purple_fill

            if key == 'ServiceMed_block_1':
                if df.loc[r_idx - start_row - 1, 'Check Type_ServiceMed_block_1'] == 'Total Number of Members':
                    if col_name in ['Count_CurrYear']:
                        # print(col_name)
                        # print(mem_cnt_val,sub_cnt_val)
                        # print('In')
                        mem_cnt_val != ser_cnt_val
                        cell.fill = purple_fill

            if key == 'ServiceMed_block_2':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_2'] == '# Members without Eligibility Record':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_3':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_3'] == '# Records with DOS within Report Period':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 100:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_3':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_3'] == '# Records with Pay Date within Report Period':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 100:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_4':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_4'] == '# Distinct Servicing Providers without Provider Table Record':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_9':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_9'] == '# Records with Pay Date before DOS':
                    # Check for both Value_PrevYear and Value_CurrYear
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_9':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_9'] == '# Records with Pay Date before To Date':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_9':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_9'] == '# Records with To Date before From Date':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_11':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_11'] == '# Records with Negative Allowed Amount':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_13':
                if df.loc[
                    r_idx - start_row - 1, 'Check Type_ServiceMed_block_13'] == '# Records with duplicate Unique Record ID':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase is None or percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_14':
                if df.loc[
                    r_idx - start_row - 1, 'Check for Length Issues_ServiceMed_block_14'] == '# Records with Member ID > 32':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase is None or percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_14':
                if df.loc[
                    r_idx - start_row - 1, 'Check for Length Issues_ServiceMed_block_14'] == '# Records Procedure Code > 5':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase is None or percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'ServiceMed_block_14':
                if df.loc[
                    r_idx - start_row - 1, 'Check for Length Issues_ServiceMed_block_14'] == '# Records with Procedure Code Modifier > 2':
                    if col_name in ['Percent_CurrYear']:
                        try:
                            percentage_increase = float(value.replace('(', '').replace(')', '').replace('%', ''))
                            if percentage_increase is None or percentage_increase != 0:
                                cell.fill = purple_fill
                        except ValueError as e:
                            print(f"Error converting Val_diff to float: {e}")

            if key == 'Member_block_18':
                if df.loc[r_idx - start_row - 1, 'Address_Member_block_18'] == '# Records with State = "XX"':
                    # print(f"Row {r_idx}, Column {c_idx}: Check Type matches")
                    if col_name in ['Record Count_PrevYear', 'Record Count_CurrYear']:
                        # print(f"Row {r_idx}, Column {c_idx}: Column name is Record Count")
                        if value is None or value == '':
                            cell.fill = purple_fill  # Highlight non-zero values

            if key == 'Member_block_8' and col_name in ['Record Count_PrevYear', 'Record Count_CurrYear']:
                if isinstance(value, (int, float)) and value != 0:
                    cell.fill = purple_fill  # Highlight non-zero values

            if key == 'Member_block_7':
                if df.loc[r_idx - start_row - 1, 'Date of Birth_Member_block_7'] == 'First DOB':
                    # print(value)
                    if col_name == ['Value_PrevYear']:
                        # print(col_name,value)
                        value == "10/01/1800"
                        cell.fill = purple_fill

                        # Function to get the last date of March for a given year

            def get_last_date_of_march(year):
                return datetime(year, 3, 31).strftime('%m/%d/%Y')

            year = datetime.now().year  # Replace with the actual year if available
            last_date_of_march = get_last_date_of_march(year)
            # print(last_date_of_march)

            if key == 'Member_block_7':
                if df.loc[r_idx - start_row - 1, 'Date of Birth_Member_block_7'] == 'Last DOB':
                    if col_name in ['Value_PrevYear', 'Value_CurrYear']:
                        # Assuming the year is part of the value or can be derived from another column
                        year = datetime.now().year  # Replace with the actual year if available
                        last_date_of_march = get_last_date_of_march(year)
                        try:
                            # Ensure value is a string before converting to datetime
                            value_date = datetime.strptime(str(value), '%Y-%m-%d')  # Adjust format as needed
                            if value_date < last_date_of_march:
                                cell.fill = purple_fill
                        except ValueError:
                            pass  # Handle the case where the value is not a valid date

            if key == 'Subscriber_block_1':
                if df.loc[r_idx - start_row - 1, 'Check Type_Subscriber_block_1'] == 'Total Number of Subscribers':
                    if col_name in ['Value_CurrYear']:
                        # print(col_name)
                        # print(mem_cnt_val,sub_cnt_val)
                        # print('In')
                        mem_cnt_val != sub_cnt_val
                        cell.fill = purple_fill

    return mismatch_count


# Save the merged DataFrames to a new Excel workbook
merged_workbook = Workbook()

# Create a summary sheet
# summary_sheet = merged_workbook.create_sheet(title="Summary")
# summary_sheet.append(["Sheet Name", "Block Name", "Mismatch Count"])

# Write each DataFrame to the corresponding sheet and update the summary
for key, df in merged_dataframes.items():
    sheet_name, block_name = key.split('_block_')
    if sheet_name not in merged_workbook.sheetnames:
        sheet = merged_workbook.create_sheet(title=sheet_name)
        start_row = 1  # Start at the first row for new sheets
    else:
        sheet = merged_workbook[sheet_name]
        start_row = sheet.max_row + 2  # Add an empty row between blocks

    mismatch_count = compare_and_color_code(sheet, df, start_row, key)
#   summary_sheet.append([sheet_name, block_name, mismatch_count])

# merged_dataframes['Member_block_8']

# Remove the default sheet created by Workbook
# if 'Sheet' in merged_workbook.sheetnames:
#   merged_workbook.remove(merged_workbook['Sheet'])

# Save the workbook
merged_workbook.save('Comparision_Res.xlsx')
print('Comparision File Ready')
